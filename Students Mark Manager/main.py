from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from db import Database
from openpyxl import Workbook
from openpyxl import *

db = Database("Students.db")
root = Tk()
root.title("Students mark recorder")
root.geometry("1366x768+0+0")
root.config(bg="#2c3e50")
root.state("zoomed")

roll = StringVar()
name = StringVar()
english = StringVar()
maths = StringVar()
chemistry = StringVar()
physics = StringVar()
computer = StringVar()


entries_frame = Frame(root, bg="#535c68")
entries_frame.pack(side=TOP, fill=X)
title = Label(
    entries_frame,
    text="Student mark manager",
    font=("Calibri", 18, "bold"),
    bg="#535c68",
    fg="white",
)
title.grid(row=0, columnspan=2, padx=5, pady=20, sticky="w")

lblroll = Label(
    entries_frame, text="Roll_no", font=("Calibri", 16), bg="#535c68", fg="white"
)
lblroll.grid(row=1, column=0, padx=5, pady=5, sticky="w")
txtroll = Entry(entries_frame, textvariable=roll, font=("Calibri", 16), width=30)
txtroll.grid(row=1, column=1, padx=5, pady=5, sticky="w")


lblName = Label(
    entries_frame, text="Name", font=("Calibri", 16), bg="#535c68", fg="white"
)
lblName.grid(row=1, column=2, padx=5, pady=5, sticky="w")
txtName = Entry(entries_frame, textvariable=name, font=("Calibri", 16), width=30)
txtName.grid(row=1, column=3, padx=5, pady=5, sticky="w")

lblEng = Label(
    entries_frame, text="Eng", font=("Calibri", 16), bg="#535c68", fg="white"
)
lblEng.grid(row=2, column=0, padx=5, pady=5, sticky="w")
txtEng = Entry(entries_frame, textvariable=english, font=("Calibri", 16), width=30)
txtEng.grid(row=2, column=1, padx=5, pady=5, sticky="w")

lblmaths = Label(
    entries_frame, text="Maths", font=("Calibri", 16), bg="#535c68", fg="white"
)
lblmaths.grid(row=2, column=2, padx=5, pady=5, sticky="w")
txtmaths = Entry(entries_frame, textvariable=maths, font=("Calibri", 16), width=30)
txtmaths.grid(row=2, column=3, padx=5, pady=5, sticky="w")

lblchemistry = Label(
    entries_frame, text="chemistry", font=("Calibri", 16), bg="#535c68", fg="white"
)
lblchemistry.grid(row=3, column=0, padx=5, pady=5, sticky="w")
txtchemistry = Entry(
    entries_frame, textvariable=chemistry, font=("Calibri", 16), width=30
)
txtchemistry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

lblphysics = Label(
    entries_frame, text="Physics", font=("Calibri", 16), bg="#535c68", fg="white"
)
lblphysics.grid(row=3, column=2, padx=5, pady=5, sticky="w")
txtphysics = Entry(entries_frame, textvariable=physics, font=("Calibri", 16), width=30)
txtphysics.grid(row=3, column=3, padx=5, pady=5, sticky="w")


lblcomputer = Label(
    entries_frame, text="computer", font=("Calibri", 16), bg="#535c68", fg="white"
)
lblcomputer.grid(row=4, column=0, padx=5, pady=5, sticky="w")
txtcomputer = Entry(
    entries_frame, textvariable=computer, font=("Calibri", 16), width=30
)
txtcomputer.grid(row=4, column=1, padx=5, sticky="w")


def getData(event):
    selected_row = tv.focus()
    data = tv.item(selected_row)
    global row
    row = data["values"]

    roll.set(row[0])
    name.set(row[1])
    english.set(row[2])
    maths.set(row[3])
    chemistry.set(row[4])
    physics.set(row[5])
    computer.set(row[6])


def dispalyAll():
    tv.delete(*tv.get_children())
    for row in db.fetch():
        tv.insert("", END, values=row)


def add_employee():
    if (
        txtroll.get() == ""
        or txtName.get() == ""
        or txtEng.get() == ""
        or txtmaths.get() == ""
        or txtchemistry.get() == ""
        or txtphysics.get() == ""
        or txtcomputer.get() == ""
    ):
        messagebox.showerror("Erorr in Input", "Please Fill All the Details")
        return
    db.insert(
        int(txtroll.get()),
        txtName.get(),
        int(txtEng.get()),
        int(txtmaths.get()),
        int(txtchemistry.get()),
        int(txtphysics.get()),
        int(txtcomputer.get()),
    )
    messagebox.showinfo("Success", "Record Inserted")
    clearAll()
    dispalyAll()


def update_employee():
    if (
        txtroll.get() == ""
        or txtName.get() == ""
        or txtEng.get() == ""
        or txtmaths.get() == ""
        or txtchemistry.get() == ""
        or txtphysics.get() == ""
        or txtcomputer.get() == ""
    ):
        messagebox.showerror("Erorr in Input", "Please Fill All the Details")
        return
    db.update(
        int(txtroll.get()),
        txtName.get(),
        int(txtEng.get()),
        int(txtmaths.get()),
        int(txtchemistry.get()),
        int(txtphysics.get()),
        int(txtcomputer.get()),
    )
    messagebox.showinfo("Success", "Record Update")
    clearAll()
    dispalyAll()


def delete_employee():
    db.remove(row[0])
    clearAll()
    dispalyAll()


def clearAll():
    roll.set("")
    name.set("")
    english.set("")
    maths.set("")
    chemistry.set("")
    physics.set("")
    computer.set("")


def excel():
    try:
        wb = load_workbook(filename="Students.xlsx")
    except:
        wb = Workbook()
    sheet = wb.active

    res = db.fetch()
    i = 0
    for row in res:
        i += 1
        j = 1
        for col in row:
            cell = sheet.cell(row=i, column=j)
            cell.value = col
            j += 1

    wb.save(filename="Students.xlsx")


btn_frame = Frame(entries_frame, bg="#535c68")
btn_frame.grid(row=6, column=0, columnspan=4, padx=5, pady=5, sticky="w")
btnAdd = Button(
    btn_frame,
    command=add_employee,
    text="Add Details",
    width=15,
    font=("Calibri", 16, "bold"),
    fg="white",
    bg="#16a085",
    bd=0,
).grid(row=0, column=0)
btnEdit = Button(
    btn_frame,
    command=update_employee,
    text="Update Details",
    width=15,
    font=("Calibri", 16, "bold"),
    fg="white",
    bg="#2980b9",
    bd=0,
).grid(row=0, column=1, padx=5)
btnDelete = Button(
    btn_frame,
    command=delete_employee,
    text="Delete Details",
    width=15,
    font=("Calibri", 16, "bold"),
    fg="white",
    bg="#c0392b",
    bd=0,
).grid(row=0, column=2, padx=5)
btnClear = Button(
    btn_frame,
    command=clearAll,
    text="Clear Details",
    width=15,
    font=("Calibri", 16, "bold"),
    fg="white",
    bg="#f39c12",
    bd=0,
).grid(row=0, column=3, padx=5)

btnexcel = Button(
    btn_frame,
    command=excel,
    text="Get Excel",
    width=15,
    font=("Calibri", 16, "bold"),
    fg="white",
    bg="#bf8040",
    bd=0,
).grid(row=0, column=4, padx=5)


tree_frame = Frame(root, bg="#ecf0f1")
tree_frame.place(x=0, y=435, width=1366, height=480)
style = ttk.Style()
style.configure("mystyle.Treeview", font=("Calibri", 18), rowheight=50)
style.configure("mystyle.Treeview.Heading", font=("Calibri", 18))
tv = ttk.Treeview(tree_frame, columns=(1, 2, 3, 4, 5, 6, 7), style="mystyle.Treeview")
tv.heading("1", text="Roll_no")
tv.column("1", width=7)
tv.heading("2", text="Name")
tv.heading("3", text="English")
tv.column("3", width=7)
tv.heading("4", text="Maths")
tv.column("4", width=7)
tv.heading("5", text="chemistry")
tv.column("5", width=7)
tv.heading("6", text="physics")
tv.column("6", width=7)
tv.heading("7", text="computer")
tv.column("7", width=7)
tv["show"] = "headings"
tv.bind("<ButtonRelease-1>", getData)
tv.pack(fill=X)

dispalyAll()
root.mainloop()
